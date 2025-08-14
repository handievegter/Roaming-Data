import streamlit as st 
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import random
from typing import cast
from openpyxl.worksheet.worksheet import Worksheet

def round2(series_or_value):
    return pd.to_numeric(series_or_value, errors="coerce").round(2)

def floor2(series_or_value):
    return np.floor(pd.to_numeric(series_or_value, errors="coerce") * 100) / 100

st.set_page_config(page_title="Roaming Data Cost Aggregator", page_icon="💸")


# --- Helper function: spacing ---
def add_vertical_space(lines=1):
    st.markdown("<br>" * lines, unsafe_allow_html=True)

# --- Data cleaning function ---
def clean_roaming_data(file, cut_off=20):
    xls = pd.ExcelFile(file, engine="openpyxl")
    sheet_name = xls.sheet_names[0]
    df = xls.parse(sheet_name, skiprows=5)

    # Ensure the standardized 7 columns are present and aligned
    if df.shape[1] < 7:
        raise ValueError(f"Expected at least 7 columns after skipping headers; got {df.shape[1]}. Please verify the input format.")
    df = df.iloc[:, :7]
    df.columns = [
        "MSISDN", "Transporter", "VehicleReg",
        "CallsRoaming", "CallsData", "TotalExclVAT", "Old Total"
    ]
    # Ensure MSISDN is always string and stripped
    df["MSISDN"] = df["MSISDN"].fillna("").astype(str).str.strip()

    numeric_cols = ["CallsRoaming", "CallsData", "TotalExclVAT", "Old Total"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).round(2)

    df["New Total"] = df["Old Total"]
    # Normalize transporter and derive grouping key without trailing "BUP"
    df["Transporter"] = df["Transporter"].fillna("").astype(str).str.strip()
    df["TransporterGroup"] = df["Transporter"].str.replace(r"\s*BUP$", "", regex=True).str.strip()
    # df["Status"] = ""

    # ------------------------------------------------------------
    # Conditional pre-aggregation:
    # Only merge rows when there are multiple VehicleReg values that
    # share the same base (with BUP removed) AND at least one of them
    # includes a trailing "BUP".
    # ------------------------------------------------------------
    df["VehicleReg"] = df["VehicleReg"].fillna("").astype(str).str.strip()
    df["VehicleRegBase"] = df["VehicleReg"].str.replace(r"\s*BUP$", "", regex=True).str.strip()
    df["HasBUP"] = df["VehicleReg"].str.contains(r"\s*BUP$", regex=True)
    
    rows = []
    numeric_cols = ["CallsRoaming", "CallsData", "TotalExclVAT", "Old Total", "New Total"]
    
    for (tgrp, vbase), g in df.groupby(["TransporterGroup", "VehicleRegBase"], as_index=False):
        if g.shape[0] >= 2 and g["HasBUP"].any():
            # Prefer the MSISDN from the non-BUP vehicle (exact base reg)
            msisdn_series_base = (
                g.loc[~g["HasBUP"], "MSISDN"].astype(str).str.strip()
            )
            msisdn_base = msisdn_series_base[msisdn_series_base.ne("")].iloc[0] if not msisdn_series_base.empty and msisdn_series_base.ne("").any() else ""

            # Fallback: first non-empty MSISDN from the whole group
            if msisdn_base == "":
                msisdn_series_any = g["MSISDN"].astype(str).str.strip()
                non_empty = msisdn_series_any[msisdn_series_any.ne("")]
                msisdn_base = non_empty.iloc[0] if not non_empty.empty else ""

            # Merge into a single row using the base reg, summing numeric fields
            summed = {col: round(float(g[col].sum()), 2) for col in numeric_cols}
            rows.append({
                "MSISDN": str(msisdn_base),
                "Transporter": tgrp,               # normalize to group label
                "VehicleReg": vbase,               # use base reg (no BUP)
                "CallsRoaming": summed["CallsRoaming"],
                "CallsData": summed["CallsData"],
                "TotalExclVAT": summed["TotalExclVAT"],
                "Old Total": summed["Old Total"],
                "New Total": summed["Old Total"],
                # "Status": "",
                "TransporterGroup": tgrp
            })
        else:
            # Keep original rows (no merge). Normalize Transporter to group label.
            for _, r in g.iterrows():
                rows.append({
                    "MSISDN": str(r.get("MSISDN", "")),
                    "Transporter": tgrp,             # normalize to group label
                    "VehicleReg": r["VehicleReg"],   # keep original (may include BUP)
                    "CallsRoaming": round(float(r["CallsRoaming"]), 2),
                    "CallsData": round(float(r["CallsData"]), 2),
                    "TotalExclVAT": round(float(r["TotalExclVAT"]), 2),
                    "Old Total": round(float(r["Old Total"]), 2),
                    "New Total": round(float(r["New Total"]), 2),
                    # "Status": "",
                    "TransporterGroup": tgrp
                })
    
    df = pd.DataFrame(rows, columns=[
        "MSISDN", "Transporter", "VehicleReg",
        "CallsRoaming", "CallsData", "TotalExclVAT",
        "Old Total", "New Total", "TransporterGroup"
    ])

    result_rows = []
    totals_rows = []

    for transporter, group in df.groupby("TransporterGroup"):
        group = group.copy().sort_values(by="VehicleReg").reset_index(drop=True)

        is_small = (group["Old Total"] < cut_off) & (group["Old Total"] > 0)
        is_large = (group["Old Total"] >= cut_off)

        if is_large.any():
            small_idxs = group[is_small].index.tolist()
            if small_idxs:
                candidate_idxs = group[is_large].index.tolist()
                for idx in small_idxs:
                    # Choose a target from large candidates (excluding self if applicable)
                    valid_targets = [i for i in candidate_idxs if i != idx]
                    if not valid_targets:
                        continue
                    target_idx = random.choice(valid_targets)
                    group.at[target_idx, "New Total"] += group.at[idx, "Old Total"]
                    group.at[idx, "New Total"] = 0
        else:
            if group.empty:
                continue
            total_sum = group["Old Total"].sum()
            collector_candidates = group.index.tolist()
            if not collector_candidates:
                continue
            collector_idx = random.choice(collector_candidates)
            for idx in group.index:
                if idx == collector_idx:
                    group.at[idx, "New Total"] = total_sum
                else:
                    group.at[idx, "New Total"] = 0

        group = group.drop(columns="TransporterGroup")

        # Floor per‑row New Total to 2 decimals BEFORE computing totals
        group["New Total"] = pd.to_numeric(group["New Total"], errors="coerce")
        group["New Total"] = np.floor(group["New Total"] * 100) / 100

        for c in ["CallsRoaming", "CallsData", "TotalExclVAT", "Old Total"]:
            group[c] = pd.to_numeric(group[c], errors="coerce").round(2)

        # Compute and store totals for this transporter group based on floored per‑row values
        total_old = group["Old Total"].sum()
        total_new = group["New Total"].sum()
        totals_rows.append({
            "Transporter": transporter,
            "Old Total": total_old,
            "New Total": total_new
        })

        # Append the group's rows
        result_rows.append(group)

        # Append a visible Grand Total row to the main sheet
        total_row = {
            "MSISDN": "",
            "Transporter": f"{transporter} - Grand Total",
            "VehicleReg": "",
            "CallsRoaming": "",
            "CallsData": "",
            "TotalExclVAT": "",
            "Old Total": total_old,
            "New Total": total_new,
            # "Status": "total"
        }
        result_rows.append(pd.DataFrame([total_row]))

        # Add two spacer rows, with New Total set to NaN so Excel ignores them in sums
        empty_row = pd.DataFrame([[""] * len(group.columns)] * 2, columns=group.columns)
        empty_row["New Total"] = np.nan
        result_rows.append(empty_row)

    final_df = pd.concat(result_rows, ignore_index=True)
    final_df["New Total"] = pd.to_numeric(final_df["New Total"], errors="coerce")
    final_df["New Total"] = np.floor(final_df["New Total"] * 100) / 100

    for c in ["CallsRoaming", "CallsData", "TotalExclVAT", "Old Total"]:
        final_df[c] = pd.to_numeric(final_df[c], errors="coerce").round(2)

    # Blank out "New Total" for spacer rows between transporters
    spacer_mask = final_df["Transporter"].astype(str).str.strip().eq("")
    final_df.loc[spacer_mask, "New Total"] = ""

    # Do not floor or round "Old Total" anywhere; leave as-is.

    return final_df

# --- Excel export with styling ---
def to_excel(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Processed')

    output.seek(0)
    wb = load_workbook(output)
    ws = cast(Worksheet, wb.active)
    if ws is None:
        raise ValueError("No active worksheet found in the workbook.")

    # Map headers to column indices
    header_map = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    msisdn_idx = header_map.get("MSISDN")
    if msisdn_idx:
        for col in ws.iter_cols(min_col=msisdn_idx, max_col=msisdn_idx, min_row=2, max_row=ws.max_row):
            for c in col:
                c.number_format = "@"

    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color="00DDDDDD", end_color="00DDDDDD", fill_type="solid")

    # Style rows where Transporter ends with " - Grand Total"
    transporter_idx = header_map.get("Transporter")
    if transporter_idx:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            tval = row[transporter_idx - 1].value
            if isinstance(tval, str) and tval.endswith(" - Grand Total"):
                for cell in row:
                    cell.font = bold_font
                    cell.fill = grey_fill

    # Apply number formats to specific numeric columns (2 decimals)
    for col_name in ["Old Total", "New Total", "TotalExclVAT", "CallsRoaming", "CallsData"]:
      col_idx = header_map.get(col_name)
      if col_idx:
          for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
              for c in cell:
                  if isinstance(c.value, (int, float)):
                      c.number_format = "0.00"

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_length:
                    max_length = len(val)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    styled_output = BytesIO()
    wb.save(styled_output)
    styled_output.seek(0)
    return styled_output

# --- App layout ---
left, center, right = st.columns([0.5, 100, 0.5])

with center:
    st.title("💸Roaming Cost Aggregator💸")

    add_vertical_space(1)

    cut_off = st.number_input(
        "Cut-Off Value",
        min_value=0,
        value=10,
        step=1,
        help="All values below this will be merged into larger ones or consolidated into one."
    )

    uploaded_file = st.file_uploader("Upload raw Excel file", type=["xlsx"])

    if uploaded_file:
        try:
            with st.spinner("Processing file..."):
                df_cleaned = clean_roaming_data(uploaded_file, cut_off)
                download_file = to_excel(df_cleaned)

            st.success("File processed successfully. Download it below:")

            st.download_button(
                label="🚀 Download cleaned Excel file",
                data=download_file,
                file_name="processed_roaming_cost.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"An error occurred: {e}")
